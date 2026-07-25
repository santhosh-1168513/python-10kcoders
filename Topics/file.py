open("example.txt", "x")


# writing to the file
f = open("example.txt", "w")
f.write("Hello, World!")
f.close()


f = open("example.txt", "r")
reading = f.read()
print(reading)
print("FILE IS CLOSED: ", f.closed)
# false means the file is still open, true means the file is closed
f.close()
print("FILE IS CLOSED: ", f.closed)


# to append to the file, we can use the "a" mode
f = open("example.txt", "a")
f.write("\n python is easy to learn.")
f.close()
# to read the file again, we can use the "r" mode
f = open("example.txt", "r")
reading = f.read()
print(reading)
f.close()


# to read the file line by line, we can use the "readline()" method
f = open("example.txt", "r")
reading = f.readline()
readings = f.readline()
print(reading)
print(readings)
f.close()


open("pythonhistory.txt", "x") # to create a new file
f = open("pythonhistory.txt", "w") # to write to the file
f.write("""Python is a high-level, interpreted, and general-purpose programming language. 
        It was created by Guido van Rossum and first released in 1991. 
        Python's design philosophy emphasizes code readability with its notable use of significant whitespace. 
        It supports multiple programming paradigms, including structured (particularly procedural), object-oriented, and functional programming. Python is dynamically typed and garbage-collected. 
        It has a large and comprehensive standard library. """)
f.close()


f = open("pythonhistory.txt", "r") # to read the file
# readings = f.readline() # to read the file line by line
# reading = f.read() # to read the entire file
# print(reading)
# print(readings)

# print(f.readline())
# print(f.readline())

for i in range(4):
    print(f.readline())

f.close()


# difference between the readline() and readlines() methods
# readline() reads a single line from the file 
# readlines() reads all the lines from the file and returns them as a list of strings.


import os
# os.rename("pythonhistory.txt", "pythonhistory1.txt") # to rename the file
# os.remove("pythonhistory1.txt") # to delete the file

# method 2
# with is a context manager that automatically closes the file after the block of code is executed
with open("pythonhistory1.txt", "r") as f:
    print(f.read())
    
with open("pythonhistory1.txt", "r") as f:
    lines = f.readlines()

with open("pythonhistory1.txt", "w") as f:
    for line in lines:
        if line.strip()!="": # to remove empty lines
            f.write(line)
print(lines)


# to remove a specific line from the file
removeline = "It was created by Guido van Rossum and first released in 1991."
with open("pythonhistory1.txt", "r") as f:
    lines = f.readlines()

with open("pythonhistory1.txt", "w") as f:
    for line in lines:
        if line.strip()!= removeline:  # to remove the specified line
            f.write(line)
print(lines)


l = [1,2,3,4,5]
for i in l:
    print(i)
# enumerate() is a built-in function that returns an enumerate object.
l = [1,2,3,4,5]
for i, j in enumerate(l):
    print(f"index values: {i}, list value: {j}")

fruits = ["apple", "banana", "cherry"]
for ind, fruit in enumerate(fruits):
    print(f"index values: {ind}, list value: {fruit}")


l = [1,2,3,4,5]
for i, j in enumerate(l, start=1): # to start the index from 1 instead of 0
    print(f"index values: {i}, list value: {j}")




fruits = ["apple", "banana", "cherry"]
with open("fruits.txt", "w") as f:
    for file in fruits:
        f.write(file + "\n") # to write each fruit in a new line

delete = 3
with open("fruits.txt", "r") as f:
    lines = f.readlines()

with open("fruits.txt", "w") as f:
    for indusnum, values in enumerate(lines):
        if indusnum != delete:
            f.write(values)


with open("fruits.txt", "r") as f:
    print(f.read())


###########################################################
import csv # to read and write csv files
# open("students.csv", "w") # to create a new csv file

with open("students.csv", "w") as f:
    writer = csv.writer(f)
    writer.writerow(["Name", "Age", "Grade"])
    writer.writerow(["santhosh", 20 , "A"])
    writer.writerow(["sai", 21 , "B"])
    writer.writerow(["kumar", 21 , "B"])

# to read the entire csv file
with open("students.csv", "r") as f:
    reader = csv.reader(f)
    for line in reader:
        print(line)


# to read the csv file line by line
with open("students.csv", "r") as f:
    print(f.readline())

with open("students.csv", "r") as f:
    lines = f.readdlines()
    print(lines[2]) # to read the 3rd line of the csv file


import csv
data = [["siva", 22, "C"], ["raju", 23, "D"]]
with open("students.csv", "a", newline="") as f:
    writer = csv.writer(f)
    writer.writerows(data) # to write multiple rows to the csv file

import csv
with open("students.csv", "r") as f:
    reader = csv.reader(f)
    for line in reader:
        print(line)


# to create a new csv file 
import csv
with open("students1.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(["Name", "Age", "cources"])
    writer.writerow(["santhoshkumar", 20 , "python"])
    writer.writerow(["sai", 21 , "java"])
    writer.writerow(["kumar", 21 , "c++"])
    writer.writerow(["siva", 22 , "javascript"])

with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    for line in reader:
        print(line)



# to count the number of students in the csv file
import csv
count = 0
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    next(reader) # to skip the header row
    for line in reader:
        count += 1
        print(line)
    print("Total number of students: ", count)


# to read specific column from the csv file
import csv
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    next(reader) # to skip the header row
    for column in reader:
        print(column[0]) # to read the first column (Name) from the csv file


# to read specific columns from the csv file 
import csv
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    next(reader) # to skip the header row
    for column in reader:
        print(column[0:2]) # to read the first two columns (Name and Age) from the csv file


# to read specific name from the csv file
import csv
search_name = "siva"
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    next(reader) # to skip the header row
    for column in reader:
        if column[0] == search_name:
            print(column)


# to show the age more than 21 from the csv file
import csv
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    next(reader)
    for column in reader:
        if int(column[1]) > 21:
            print(column)

# upadate the specific name in the csv file
rows = []
import csv
with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    for r in reader:
        if r[0] == "siva":
            r[1] = "25" # update the age to 25  
        rows.append(r)

with open("students1.csv", "w", newline="") as f:
    writer = csv.writer(f)
    writer.writerows(rows)

with open("students1.csv", "r") as f:
    reader = csv.reader(f)
    for line in reader:
        print(line)


############################################################################
# openpyxl is a Python library to read and write Excel files (xlsx/xlsm/xltx/xltm files).
# pip install openpyxl
# workbook is a collection of worksheets, and worksheet is a single sheet in the workbook.

import openpyxl  # to read and write excel files
# from openpyxl import Workbook
wb = openpyxl.Workbook()  # to create a new workbook
ws = wb.active  # to get the active sheet
ws.title = "Studentdetails"  # to rename the sheet
wb.save("students.xlsx")  # to save the workbook


import openpyxl
wb = openpyxl.load_workbook("students.xlsx")  # to load the workbook
ws = wb["Studentdetails"]  # to get the sheet by name
ws["A1"] = "STUDENTID"
ws["B1"] = "STUDENTNAME"
ws["C1"] = "MARKS"
wb.save("students.xlsx")  # to save the workbook

# multiple rows of data to the excel sheet
import openpyxl
data = [[1, "santhosh", 90], 
        [2, "sai", 80], 
        [3, "kumar", 70], 
        [4, "siva", 60]
        ]

for row in data:
    ws.append(row)  # to append the data to the sheet
wb.save("students.xlsx")  # to save the workbook

ws = openpyxl.load_workbook("students.xlsx") # to load the workbook and get the sheet by name
ws =wb.active

# iterate through the rows and get the values only
import openpyxl
for row in ws.iter_rows(values_only=True):
    print(row)

print(ws["A2"].value)
print(ws["C3"].value)

ws["B3"] = "kumar" # to update the value in the cell
wb.save("students.xlsx") # to save the workbook

ws =openpyxl.load_workbook("students.xlsx") # to load the workbook and get the sheet by name
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)

for row in ws.iter_rows(values_only=True):
        print(row)

# delete a specific row from the excel sheet
ws.delete_rows(2) # to delete the 2nd row
wb.save("students.xlsx") # to save the workbook

for row in ws.iter_rows(values_only=True):
    print(row)


# 
wb.create_chartsheet("employedetails") # to create a new chart sheet


# to create a new workbook and add a new sheet to it
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "company" 
wb.save("orders.xlsx")

# print(wb.sheetnames) # to get the sheet names in the workbook
# print(wb.active) # to get the active sheet in the workbook

wb = openpyxl.load_workbook("offer.xlsx") 
ws = wb["company"]
ws["A1"] = "orderid"
ws["B1"] = "ordername"
wb.save("offer.xlsx")

data = [[1, "santhosh"],
        [2, "vani"],
        [3, "viswanath"],
        [4, "siva"]
        ]
for row in data:
    ws.append(row)
wb.save("offer.xlsx")
# read the data from the excel sheet
from openpyxl import load_workbook

wb = load_workbook("Company.xlsx")
ws = wb["Orders"]

for row in ws.iter_rows(values_only=True):
    print(row)



import openpyxl

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "company"

# Headings
ws["A1"] = "orderid"
ws["B1"] = "ordername"

# Data
data = [
    [1, "santhosh"],
    [2, "vani"],
    [3, "viswanath"],
    [4, "siva"]
]

for row in data:
    ws.append(row)

# Save workbook
wb.save("offer.xlsx")

# Read workbook
wb = openpyxl.load_workbook("offer.xlsx")
ws = wb["company"]

for row in ws.iter_rows(values_only=True):
    print(row)